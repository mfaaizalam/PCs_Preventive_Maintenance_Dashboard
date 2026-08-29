"""
Builds the "print"/export .xlsx for the preventive-maintenance checklist.

Two shapes, both for one frequency + one period:
  - single PC   (computer_id given): one sheet, that PC's tasks.
  - whole lab   (computer_id=None):  one sheet, every enrolled PC's tasks,
                grouped under their lab_section header - the same grouping
                the original master-list Excel used (e.g. "CAED PCS").

Row status (Completed / Pending / Overdue) is computed with the same
rules the dashboard UI uses (see dashboard/src/utils/period.js -
isPeriodElapsed / rowState). period_label is treated as an opaque string
by the rest of the backend, but the export needs to know whether a period
has fully elapsed, so this module re-implements that same small parser
in Python. If you change the label convention on the frontend, mirror
the change here too.
"""

from __future__ import annotations

import calendar
import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.models.computer import Computer
from app.services import maintenance_service

# ---------------------------------------------------------------- styling

HEADER_FILL = PatternFill("solid", fgColor="0F766E")  # teal, matches dashboard brand color
SECTION_FILL = PatternFill("solid", fgColor="E7F6EF")
COMPLETED_FILL = PatternFill("solid", fgColor="E7F6EF")
OVERDUE_FILL = PatternFill("solid", fgColor="FCEAEA")
PENDING_FILL = PatternFill("solid", fgColor="FDF2E1")

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(name="Arial", bold=True, size=11, color="0F766E")
TITLE_FONT = Font(name="Arial", bold=True, size=14)
SUBTITLE_FONT = Font(name="Arial", size=10, color="666666")
BODY_FONT = Font(name="Arial", size=10)

COLUMNS = ["S.No", "PC Name", "Specification", "Task", "Responsible", "Status", "Completed By", "Completed At"]
COL_WIDTHS = [7, 20, 32, 34, 16, 12, 18, 20]


# ------------------------------------------------------------ period math

def _period_end_date(frequency: str, label: str) -> datetime | None:
    """Mirrors dashboard/src/utils/period.js:periodEndDate."""
    try:
        if frequency == "biweekly":
            year, month, week = label.split("-")
            year, month = int(year), int(month)
            is_second_half = week in ("W2", "2")
            last_day = calendar.monthrange(year, month)[1]
            day = last_day if is_second_half else 15
            return datetime(year, month, day, 23, 59, 59, tzinfo=timezone.utc)
        if frequency in ("monthly", "custom"):
            year, month = label.split("-")
            year, month = int(year), int(month)
            last_day = calendar.monthrange(year, month)[1]
            return datetime(year, month, last_day, 23, 59, 59, tzinfo=timezone.utc)
        if frequency == "half_yearly":
            year, half = label.split("-")
            year = int(year)
            end_month = 6 if half == "H1" else 12
            last_day = calendar.monthrange(year, end_month)[1]
            return datetime(year, end_month, last_day, 23, 59, 59, tzinfo=timezone.utc)
    except (ValueError, IndexError):
        return None
    return None


def _row_status(completed: bool, frequency: str, period_label: str) -> str:
    if completed:
        return "Completed"
    end = _period_end_date(frequency, period_label)
    if end and datetime.now(timezone.utc) > end:
        return "Overdue"
    return "Pending"


_STATUS_FILL = {"Completed": COMPLETED_FILL, "Overdue": OVERDUE_FILL, "Pending": PENDING_FILL}


def _spec_string(computer: Computer) -> str:
    """Mirrors the "CPU · RAM · storage" line shown on the dashboard's
    maintenance detail page, e.g. "Intel(R) Core(TM) i7-10610U CPU @
    1.80GHz · 16 GB RAM · 476.1 GB storage"."""
    parts = []
    if computer.cpu_model:
        parts.append(computer.cpu_model)
    if computer.ram_total_gb:
        parts.append(f"{computer.ram_total_gb:g} GB RAM")
    if computer.disk_total_gb:
        parts.append(f"{computer.disk_total_gb:g} GB storage")
    return " · ".join(parts) if parts else "-"


# ------------------------------------------------------------------ sheet

def _write_header(ws: Worksheet, title: str, subtitle: str, start_row: int) -> int:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(COLUMNS))
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.font = TITLE_FONT
    ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=len(COLUMNS))
    sub_cell = ws.cell(row=start_row + 1, column=1, value=subtitle)
    sub_cell.font = SUBTITLE_FONT
    return start_row + 3


def _write_column_headers(ws: Worksheet, row: int) -> int:
    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    return row + 1


def _write_computer_block(
    ws: Worksheet,
    row: int,
    computer: Computer,
    checklist: list[dict],
    frequency: str,
    period_label: str,
) -> int:
    """Writes one PC's task rows and merges its S.No / PC Name /
    Specification cells vertically across however many task rows it has -
    this is what makes the sheet show each PC's identity once instead of
    repeating it on every task line (matches the reference master list)."""
    if not checklist:
        return row

    start_row = row
    spec = _spec_string(computer)

    # s_no is admin-assigned and usually never set, which left this
    # column blank in the exported sheet. Fall back to asset_id (which
    # IS populated automatically from the hostname on check-in) so the
    # column always shows something useful. s_no still wins when
    # someone has deliberately set it to match a paper master list.
    display_no = computer.s_no if computer.s_no is not None else (computer.asset_id or "")

    for i, item in enumerate(checklist):
        status = _row_status(item["completed"], frequency, period_label)
        values = [
            display_no if i == 0 else "",
            computer.hostname if i == 0 else "",
            spec if i == 0 else "",
            item["task_name"],
            item["responsible_person"] or "-",
            status,
            item["completed_by"] or "-",
            item["completed_at"].strftime("%Y-%m-%d %H:%M") if item["completed_at"] else "-",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = BODY_FONT
            if col_idx <= 3:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx == 6:  # Status column
                cell.fill = _STATUS_FILL[status]
        row += 1

    end_row = row - 1
    if end_row > start_row:
        for col_idx in (1, 2, 3):  # S.No, PC Name, Specification
            ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)

    return row


def _apply_column_widths(ws: Worksheet) -> None:
    for idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


# ------------------------------------------------------------------ build

def build_single_computer_workbook(
    db: Session, computer: Computer, frequency: str, period_label: str
) -> io.BytesIO:
    checklist = maintenance_service.get_checklist(db, computer.id, period_label, frequency)

    wb = Workbook()
    ws = wb.active
    ws.title = frequency.replace("_", " ").title()[:31]

    row = _write_header(
        ws,
        f"{computer.hostname} — {frequency.replace('_', ' ').title()} Maintenance Checklist",
        f"Period: {period_label}    ·    Lab: {computer.lab_section or '-'}    ·    "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        1,
    )
    row = _write_column_headers(ws, row)
    _write_computer_block(ws, row, computer, checklist, frequency, period_label)
    ws.freeze_panes = f"A{row}"
    _apply_column_widths(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_whole_lab_workbook(db: Session, frequency: str, period_label: str) -> io.BytesIO:
    computers = maintenance_service.list_computers_for_maintenance(db)

    wb = Workbook()
    ws = wb.active
    ws.title = frequency.replace("_", " ").title()[:31]

    row = _write_header(
        ws,
        f"Master List — {frequency.replace('_', ' ').title()} Preventive Maintenance Checklist",
        f"Period: {period_label}    ·    {len(computers)} PCs    ·    "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        1,
    )
    row = _write_column_headers(ws, row)

    current_section = object()  # sentinel so the first section always prints its header
    for computer in computers:
        section = computer.lab_section or "Unassigned"
        if section != current_section:
            current_section = section
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
            section_cell = ws.cell(row=row, column=1, value=section)
            section_cell.font = SECTION_FONT
            section_cell.fill = SECTION_FILL
            row += 1

        checklist = maintenance_service.get_checklist(db, computer.id, period_label, frequency)
        row = _write_computer_block(ws, row, computer, checklist, frequency, period_label)

    ws.freeze_panes = "A1"
    _apply_column_widths(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_filename(frequency: str, period_label: str, computer: Computer | None) -> str:
    safe_period = period_label.replace("/", "-")
    scope = computer.hostname if computer else "master-list"
    return f"maintenance_{frequency}_{scope}_{safe_period}.xlsx"